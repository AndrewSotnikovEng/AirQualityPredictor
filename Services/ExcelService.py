from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import time
from plistlib import InvalidFileException
import re
from Services.BackupManager import BackupManager

class ExcelService:
    path_to_file = "data.xlsx"
    wb = None  # Initialize the workbook attribute
    headers = [
        "Updated Timestamp", "Fetched Timestamp","PM1", "PM 2.5", "PM 10",
        "NO", "NO2", "CO2", "O3", "H2S", "SO2", "CH2O",
        "Temperature", "Humidity", "Pressure", "HECA Temperature",
        "HECA Humidity", "Timestamp"
    ]
    last_row = None

    @staticmethod
    def sanitize_sheet_name(name):
        name = re.sub(r'[\[\]\*\/\\:\?]', '_', name)
        return name[:31]

    @staticmethod
    def prepare_excel_file(location_names):
        """
        Ensure that the Excel file exists, restoring from backup if necessary,
        or creating a new file if neither exists.
        """
        if not os.path.exists(ExcelService.path_to_file):
            print(f"{ExcelService.path_to_file} does not exist. Attempting to restore from backup.")
            BackupManager.restore_latest_backup()

        if not os.path.exists(ExcelService.path_to_file):
            print(f"Restoration failed or no backups found. Creating a new {ExcelService.path_to_file}.")
            ExcelService.create_file(location_names)

        else:
            # Load the existing workbook
            ExcelService.create_file(location_names)

    @staticmethod
    def create_file(location_names):
        try:
            ExcelService.wb = load_workbook(ExcelService.path_to_file)
        except (FileNotFoundError, InvalidFileException):
            ExcelService.wb = Workbook()
            default_sheet = ExcelService.wb.active
            ExcelService.wb.remove(default_sheet)

        for location in location_names:
            sheet_name = ExcelService.sanitize_sheet_name(location)
            if sheet_name not in ExcelService.wb.sheetnames:
                sheet = ExcelService.wb.create_sheet(title=sheet_name)
                for col_num, header in enumerate(ExcelService.headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)

        ExcelService.wb.save(ExcelService.path_to_file)

    @staticmethod
    def get_sheet(tab_name):
        if ExcelService.wb is None:
            raise AttributeError("Workbook is not initialized. Call create_file() first.")

        sheet_name = ExcelService.sanitize_sheet_name(tab_name)
        if sheet_name in ExcelService.wb.sheetnames:
            ExcelService.sheet = ExcelService.wb[sheet_name]
        else:
            raise ValueError(f"Sheet {sheet_name} does not exist.")

    @staticmethod
    def get_max_row():
        max_row = 1
        for sheet in ExcelService.wb.worksheets:
            max_row = max(max_row, sheet.max_row)
        ExcelService.last_row = max_row + 1

    @staticmethod
    def write_to_file(aiq_items):
        if ExcelService.wb is None:
            raise AttributeError("Workbook is not initialized. Call create_file() first.")
        if (ExcelService.last_row == None):
            ExcelService.get_max_row()

        for aiq_item in aiq_items:
            sheet_name = ExcelService.sanitize_sheet_name(aiq_item.station_name)
            if sheet_name in ExcelService.wb.sheetnames:
                sheet = ExcelService.wb[sheet_name]

                current_date = time.strftime("%d-%m-%Y %H:%M:%S")
                sheet.cell(row=ExcelService.last_row, column=1, value=current_date)
                sheet.cell(row=ExcelService.last_row, column=2, value=aiq_item.date)
                sheet.cell(row=ExcelService.last_row, column=3, value=aiq_item.pm1)
                sheet.cell(row=ExcelService.last_row, column=4, value=aiq_item.pm2_5)
                sheet.cell(row=ExcelService.last_row, column=5, value=aiq_item.pm10)
                sheet.cell(row=ExcelService.last_row, column=6, value=aiq_item.no)
                sheet.cell(row=ExcelService.last_row, column=7, value=aiq_item.no2)
                sheet.cell(row=ExcelService.last_row, column=8, value=aiq_item.co2)
                sheet.cell(row=ExcelService.last_row, column=9, value=aiq_item.o3)
                sheet.cell(row=ExcelService.last_row, column=10, value=aiq_item.h2s)
                sheet.cell(row=ExcelService.last_row, column=11, value=aiq_item.so2)
                sheet.cell(row=ExcelService.last_row, column=12, value=aiq_item.ch2o)
                sheet.cell(row=ExcelService.last_row, column=13, value=aiq_item.temperature)
                sheet.cell(row=ExcelService.last_row, column=14, value=aiq_item.humidity)
                sheet.cell(row=ExcelService.last_row, column=15, value=aiq_item.pressure)
                sheet.cell(row=ExcelService.last_row, column=16, value=aiq_item.heca_temperature)
                sheet.cell(row=ExcelService.last_row, column=17, value=aiq_item.heca_humidity)
                sheet.cell(row=ExcelService.last_row, column=18, value=round(time.time()))

        ExcelService.wb.save(ExcelService.path_to_file)

    @staticmethod
    def format_tab(tab_name, font_size=12, bold=True, align_center=True):
        ExcelService.get_sheet(tab_name)

        for col_num in range(1, len(ExcelService.headers) + 1):
            cell = ExcelService.sheet.cell(row=1, column=col_num)
            cell.font = Font(size=font_size, bold=bold)
            ExcelService.sheet.column_dimensions[get_column_letter(col_num)].width = 25
            if align_center:
                cell.alignment = Alignment(horizontal='center')
