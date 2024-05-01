import pandas as pd
from datetime import datetime
import pytz
from openpyxl import load_workbook

def convert_to_utc_and_save_new(file_path, output_file, timestamp_column_name):
    # Load the Excel file with all sheets into a DataFrame dictionary
    xls = pd.read_excel(file_path, sheet_name=None)
    
    # Load the workbook to access column dimensions
    wb = load_workbook(file_path)
    
    # Create a Pandas Excel writer for the new output file using the 'openpyxl' engine
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in xls.items():
            if timestamp_column_name in df.columns:
                # Convert timestamps to UTC and add as a new column
                df['UTC Timestamp'] = df[timestamp_column_name].apply(lambda x: convert_to_utc(x))
                # Write DataFrame to the writer (but not to file yet)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Load the current sheet in the writer's workbook
                worksheet = writer.sheets[sheet_name]
                # Get column dimensions from the original workbook's sheet
                if sheet_name in wb.sheetnames:
                    original_sheet = wb[sheet_name]
                    for col in original_sheet.column_dimensions:
                        worksheet.column_dimensions[col].width = original_sheet.column_dimensions[col].width
            else:
                print(f"'{timestamp_column_name}' not found in '{sheet_name}'. Skipping this sheet.")
    
    print("Conversion complete. New file with UTC timestamps has been saved.")

def convert_to_utc(date_str):
    try:
        local_time = datetime.strptime(date_str, '%d-%m-%y %H:%M')
        timezone = pytz.timezone("Europe/Kiev")  # Adjust timezone as needed
        local_dt = timezone.localize(local_time, is_dst=None)
        utc_dt = local_dt.astimezone(pytz.utc)
        return int(utc_dt.timestamp())
    except ValueError:
        print(f"Error converting date: {date_str}. It does not match format '%d-%m-%y %H:%M'.")
        return None

# File paths
original_file_path = 'generic_output.xlsx'
new_file_path = 'updated_generic_output.xlsx'
timestamp_column_name = 'Updated Timestamp'  # Adjust column name as needed

# Call the function to convert timestamps and create a new file
convert_to_utc_and_save_new(original_file_path, new_file_path, timestamp_column_name)
