import os
from openpyxl import Workbook, load_workbook

def merge_excel_files(directory, output_file):
    # Prepare the output workbook and the summary sheet
    out_wb = Workbook()
    summary_ws = out_wb.active
    summary_ws.title = 'Summary'
    summary_ws.append(['Filename', 'Rows Copied'])
    
    # Dictionary to store data and column widths per sheet
    data_sheets = {}

    # Process each Excel file
    for filename in os.listdir(directory):
        if filename.startswith('data_') and filename.endswith('.xlsx'):
            filepath = os.path.join(directory, filename)
            wb = load_workbook(filepath)
            row_counts = {}

            # Iterate through each sheet in the workbook
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.rows)
                row_counts[sheet_name] = len(rows) - 1  # exclude header

                # If the sheet already exists in the data_sheets, append to it
                if sheet_name in data_sheets:
                    target_ws = data_sheets[sheet_name]
                else:
                    # Otherwise, create a new sheet in the output workbook
                    target_ws = out_wb.create_sheet(title=sheet_name)
                    data_sheets[sheet_name] = target_ws
                    # Copy column widths
                    for col in ws.columns:
                        target_ws.column_dimensions[col[0].column_letter].width = ws.column_dimensions[col[0].column_letter].width
                
                # Append all rows to the target worksheet, add 'Source' in the last column
                for row in rows:
                    target_ws.append([cell.value for cell in row] + [filename])

            # Append summary info for this file
            summary_ws.append([filename, sum(row_counts.values())])

    # Remove default sheet if it was not used
    if 'Sheet' in out_wb.sheetnames and out_wb['Sheet'] == out_wb.active:
        del out_wb['Sheet']

    # Save the merged workbook
    out_wb.save(output_file)

# Specify the directory and the output file name
directory = '.'  # Current directory
output_file = 'generic_output.xlsx'  # Output file in the current directory

# Execute the function
merge_excel_files(directory, output_file)
print(f"Merged Excel file created at {output_file}")
